import requests
import sys
import xmlrpc.client
import time
from datetime import datetime
import getpass
from openpyxl import load_workbook, Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font,Fill
from os import path
import os.path

#fetch current directory
dir = os.path.dirname(__file__)
current_time = datetime.now()

def fail():
    input('Press ENTER to exit')
    sys.exit()

#Normalize the URLs. Removing any potential extra slashes before and after inputs
def normalize_slashes(url):
    url = str(url)
    segments = url.split('/')
    correct_segments = []
    for segment in segments:
        if segment != '':
            correct_segments.append(segment)
    first_segment = str(correct_segments[0])
    if first_segment.find('https') == -1:
        correct_segments = ['https:'] + correct_segments
    correct_segments[0] = correct_segments[0] + '/'
    normalized_url = '/'.join(correct_segments)
    return normalized_url

# Prompt the user to enter the fileName and validate

try:
    fileName = input("Please enter the report file name>>").strip()
    wb = load_workbook(fileName)
    ws = wb.active
    #Page URLs are listed in 'C' column in the downloaded Power BI report
    first_column = ws['C']
    #List to contain links and page IDs
    links = []
    pageIDs = []
    #Grab all links leaving first two empty rows
    for x in range(len(first_column))[3:]:
        links.append(first_column[x].value)
    #Parse all links and grab the pageIDs by slicing last element
    for id in links:
        unique_id= ((id).split('='))
        pageIDs.append(unique_id[1])


except (FileNotFoundError, InvalidFileException, IndexError) as f:
    if (not path.exists(fileName)):
        print("The specified file does not exist. Provide the existing .xlsx file")
        fail()

    else:
        if ((path.splitext(fileName)[-1] == "xlsx") == False):
            print("Please provide input file in xlsx format")
            fail()

#Prompt for User credebtials and validate them
URL=input("Enter the URL of the Confluence >>").strip()
URL = normalize_slashes(URL)
if not (URL == ('https://docops-temp.ca.com') or URL == ('https://docops.ca.com')):
    print ("Invalid URL. Please enter a valid DocOps instance URL")
    fail()

#Validate user credentials. Terminate the program if they are invalid
try:
    userID=input("Enter the user ID >>").strip()
    userPwd=getpass.getpass("Enter the password >>")
    s = xmlrpc.client.ServerProxy(URL + "/rpc/xmlrpc")
    s = s.confluence1
    token = s.login(userID,userPwd)
except xmlrpc.client.Error as e:
    print("User ID or Password is invalid. Please enter valid credentials!")
    fail()

#Store the credentials in auth variable
auth = (userID, userPwd)
headers = {'X-Atlassian-Token': 'no-check','Cache-Control': 'no-cache'}

#Method to write log messages
def write_log(text, file):
    f = open(file, 'a')           # 'a' will append to an existing file if it exists
    f.write("{}\n".format(text))  # write the text to the logfile and move to next line
    return

# Create a log directory in the current path if it does not exist
current_path = os.path.dirname(os.path.realpath(__file__))
logs_dir = "logs"
output_dir = "output"
log_filename = str(current_time.strftime('PageRefresh_%d%m%Y%H%M%S.log'))
log_filepath = os.path.join(current_path, logs_dir, log_filename)

if not os.path.exists('logs'):
    # create your subdirectory
    os.mkdir(os.path.join(current_path, logs_dir))

#Loop through PageIDs and perform POST call
startTime = time.time()
print ("Force refresh process has started at "+time.strftime("%b %d %Y %H:%M:%S", time.gmtime(startTime))+ " GMT")
write_log("Force refresh process has started at "+time.strftime("%b %d %Y %H:%M:%S", time.gmtime(startTime))+ " GMT", log_filepath)

def pageRefresh(id):
    affected_link=URL+'/rest/appfusions/lingotek/latest/document/check/'+id+'?force=true'
    response = requests.post(affected_link, headers=headers, auth=auth)
    return response.text

refreshSuccess=[]
refreshFailed=[]

#Dictionary to hold Page URLs and statuses as key value pairs
myResults ={}

#Pass all PageIds to the method to force refresh the page
for id in pageIDs:
    if (pageRefresh(id)) == '{"updated":true}':
        myResults.update({URL+"/pages/viewpage.action?pageId="+id: "Success"})
        refreshSuccess.append(id)
        print("Refreshing::"+ URL +"/pages/viewpage.action?pageId="+id+"\t"+":: Success")
        write_log("Refreshing::"+ URL +"/pages/viewpage.action?pageId="+id+"\t"+":: Success", log_filepath)
    else:
        myResults.update({URL+"/pages/viewpage.action?pageId=" + id: "Failed"})
        refreshFailed.append(id)
        print("Refreshing::"+ URL +"/pages/viewpage.action?pageId="+id+"\t"+":: Failed")
        write_log("Refreshing::"+ URL +"/pages/viewpage.action?pageId="+id+"\t"+":: Failed", log_filepath)

#Working with output file
output_dir = "output"
output_file = str(current_time.strftime('Output_%d%m%Y%H%M%S.xlsx'))
output_filepath = os.path.join(current_path, output_dir, output_file)

if not os.path.exists('output'):
# create your subdirectory
    os.mkdir(os.path.join(current_path, output_dir))

outputFile = Workbook()
sheet2 = outputFile.active
c1 = sheet2.cell(row=1, column=1)
# writing values to cells
c1.value = "PAGE URL"
c1.font = Font(bold=True)
c2 = sheet2.cell(row=1, column=2)
c2.value = "Result"
c2.font = Font(bold=True)
next_row=2
for key, values in myResults.items():
     sheet2.cell(column=1, row=next_row, value=key)
     sheet2.cell(column=2, row=next_row, value=values)
     next_row += 1

outputFile.save(output_filepath)
outputFile.close()


endTime=time.time()
print("Force refresh has completed at "+time.strftime("%b %d %Y %H:%M:%S", time.gmtime(endTime))+ " GMT" +"\n" +"Please refer the log file under logs folder for more details")

print ("Number of pages successfully refreshed: ",len(refreshSuccess))
print ("Number of pages failed to refresh: ",len(refreshFailed))
write_log("Force refresh has completed at "+time.strftime("%b %d %Y %H:%M:%S", time.gmtime(endTime))+ " GMT", log_filepath)

elapsedTime = int(time.time() - startTime)
print('Total time taken to refresh pages:''{:02d}:{:02d}:{:02d}'.format(elapsedTime // 3600, (elapsedTime % 3600 // 60), elapsedTime % 60))
write_log("###DONE###", log_filepath)

input('Press ENTER to exit')

sys.exit()